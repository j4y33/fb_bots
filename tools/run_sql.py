import csv
import shutil
from tempfile import NamedTemporaryFile
from fake_useragent import UserAgent
from openpyxl import load_workbook
from browser.display_properties import DisplayProperties
from config.config import config
from database.bot_sql import BotSql
# from pyagender import PyAgender
import numpy as np
import urllib
import cv2
from urllib.request import urlopen
from actions.random_range import RR
from os import listdir
from os.path import isfile, join


def run():
    bot_sql = BotSql()
    bot_sql.create_images_table()
    agender = PyAgender()
    with open('facebook_accounts.csv') as csv_file:

        csv_reader = csv.reader(csv_file, delimiter=',')
        line_count = 0
        head = ''
        gender = ''
        for row in csv_reader:
            if row[0] != '':

                try:
                    key, value = row[1].split('https://vk.com/')
                    head = value
                    resp = urlopen(row[2])
                    image = np.asarray(bytearray(resp.read()), dtype="uint8")
                    image = cv2.imdecode(image, cv2.IMREAD_COLOR)
                    check_gender_img = agender.detect_genders_ages(image)
                    print(check_gender_img)
                    try:
                        check_gender_img[0]['gender']
                    except IndexError:
                        continue
                    if check_gender_img[0]['gender'] > 0.5:
                        print(check_gender_img[0])
                        gender = "Female"
                    else:
                        print(check_gender_img[0])
                        gender = "Male"
                    bot_sql.add_image(bot_id='None', gender=gender, profile=head, link=row[2])
                except ValueError:
                    print("Failure")
            else:
                bot_sql.add_image(bot_id='None', gender=gender, profile=head, link=row[2])


def modify_bot_vpn_region():
    import csv
    filename = 'facebook_accounts_all.csv'
    tempfile = NamedTemporaryFile(mode='w', delete=False)

    fields = ['id', 'Email / login', 'Facebook pass', 'Name', 'Link', 'VPN', 'VPN soft', 'Friends', 'Img source',
              'Email pass', 'Date of registration']

    with open(filename, 'r') as csvfile, tempfile:
        reader = csv.DictReader(csvfile, fieldnames=fields)
        writer = csv.DictWriter(tempfile, fieldnames=fields)
        for row in reader:
            if row['VPN soft'] == "Cyberghost" or row['VPN soft'] == "CyberGhost":
                print(row['VPN soft'] + ": ", row['VPN'])
                region = input("New region name: ")
                if region == "":
                    print("Skipped")
                else:
                    print('updating row', row['id'])
                    row['VPN'] = region
                    row['VPN soft'] = 'cyberghost'
            row = {'id': row['id'],
                   'Email / login': row['Email / login'],
                   'Facebook pass': row['Facebook pass'],
                   'Name': row['Name'],
                   'Link': row['Link'],
                   'VPN': row['VPN'],
                   'VPN soft': row['VPN soft'],
                   'Friends': row['Friends'],
                   'Img source': row['Img source'],
                   'Email pass': row['Email pass'],
                   'Date of registration': row['Date of registration']}
            writer.writerow(row)
    shutil.move(tempfile.name, filename)


def read_cookies():
    import csv
    filename = 'cookies.csv'

    with open(filename, 'r') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            print(row)


def add_bots(bot_sql):
    import csv
    filename = 'facebook_accounts_all_cookies.csv'
    # {'VPN soft', 'ipvanish', 'nordvpn', 'torguard', 'cyberghost'}
    fields = ['id', 'Email / login', 'Facebook pass', 'Name', 'Link', 'VPN', 'VPN soft', 'Friends', 'Img source',
              'Email pass', 'Date of registration', 'Cookie']

    with open(filename, 'r') as csvfile:
        reader = csv.DictReader(csvfile, fieldnames=fields, delimiter=";")
        for row in reader:
            print(row)


def user_agent():
    ua = UserAgent(use_cache_server=False, verify_ssl=False)
    while True:
        user_agent = ua.chrome
        if config.default_platform in user_agent:
            return user_agent
        else:
            continue


def get_cell_sheets(bot_sql):
    workbook = load_workbook('facebook_accounts_all_cookies.xlsx')
    first_sheet = workbook.sheetnames
    worksheet = workbook[first_sheet[2]]
    start = 1
    for row in worksheet.iter_rows():
        if start == 1:
            start += 1
            continue
        cookie = 'None'
        bot = []
        cookies_row = 0
        for cell in row:
            bot.append(cell.value)
            if cookies_row == 5 and not cell.comment:
                bot.append(None)
            if cell.comment:
                bot.append(cell.comment.text)
            cookies_row += 1
        vpn_login = None
        vpn_pass = None
        size = None
        if bot[13] == 'nordvpn':
            vpn_login = 'ks@goshoponline.biz'
            vpn_pass = 'Qwer0008'
        elif bot[13] == 'torguard':
            vpn_login = 'as@goshoponline.biz'
            vpn_pass = 'Qwer0008'
        elif bot[13] == 'ipvanish':
            vpn_login = 'ks@goshoponline.biz'
            vpn_pass = 'Qwer0008'
        else:
            size = DisplayProperties().get_size()
        print(bot[3])
        try:
            bot = (bot[1],
               bot[2],
               bot[13],
               bot[12],
               vpn_login,
               vpn_pass,
               bot[6],
               False,
               'active',
               bot[3].split(' ')[0],
               bot[3].split(' ')[1],
               None,
               None,
               None,
               None,
               user_agent(),
               size,
               None,
               None,
               None,
               None,
               None,
               None,
               bot[11],
               'infinity',
               None,
               0,
               0,
               0,
               None)
            print("Run add")
            bot_sql.add_bot(bot)
        except Exception as e:
            bot_sql.roll_back()
            print(e)

# 'login', 'password', 'vpn_provider', 'vpn_region', 'vpn_login', 'vpn_password', 'cookies', 'used_status',
# 'block_status', 'bot_first_name', 'bot_last_name', 'bot_gender', 'bot_birth_day', 'bot_birth_month',
# 'bot_birth_year', 'user_agent', 'screen', 'city', 'school', 'university', 'job', 'profile_picture',
# 'bot_images_list', 'creation_date', 'last_used_date', 'total_friends', 'scraped_profiles',
# 'total_actions', 'total_errors', 'proxy'
def get_image():
    bot_sql = BotSql()
    print(bot_sql.count_follow_pages('+79913813619'))


bot_sql = BotSql()
# add_bots(bot_sql)
get_cell_sheets(bot_sql)
